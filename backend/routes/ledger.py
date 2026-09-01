"""Sprint 1 · P3.6/P3.7/P3.9 — Ledger + Export utility + Budget-vs-Actual.

Ledger is projected from the vouchers collection with a running balance computed
on-fetch (opening balance = 0 by default; can be seeded via `body_budgets.opening_balance`).

Exports produce Excel (xlsx) and PDF (letterhead A4) via openpyxl + reportlab.
Budget-vs-Actual pulls annual_budget from body.annual_grant_inr (or override) and
actuals from vouchers (Payment for outflows, Receipt for inflows).
"""
import asyncio
from datetime import datetime, timezone
from io import BytesIO

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from core.infra import api_router, db
from core.shared_services import indian_fy


def _fmt_inr(v: float) -> str:
    if v is None: return "-"
    try:
        return f"₹ {float(v):,.2f}"
    except Exception:
        return str(v)


async def _resolve_body(body_id: str) -> dict:
    body = await db.bodies.find_one({"code": body_id}, {"_id": 0})
    if not body:
        raise HTTPException(404, f"Body {body_id} not found")
    return body


async def _ledger_rows(body_id: str, fiscal_cycle: str,
                       date_from: str | None = None,
                       date_to: str | None = None) -> dict:
    """Return {'body': body, 'opening': 0.0, 'rows': [...], 'totals': {...}}"""
    body = await _resolve_body(body_id)

    q: dict = {"fiscal_cycle": fiscal_cycle, "status": "Posted"}
    # Ledger scope: entries touching this body (either as owner or as counter-party grant recipient)
    q["$or"] = [
        {"body_id": body_id},
        {"linked_module": "division_grant", "cr_account": {"$regex": body_id, "$options": "i"}},
        {"particulars": {"$regex": body_id, "$options": "i"}},
    ]
    if date_from: q["date"] = {**q.get("date", {}), "$gte": date_from}
    if date_to:   q["date"] = {**q.get("date", {}), "$lte": date_to}

    vouchers = await db.vouchers.find(q, {"_id": 0}).sort("date", 1).to_list(5000)

    # For an owner body: Receipt=Debit to bank (Cr side), Payment=Credit to bank
    # We keep it simple: outflow (Payment) reduces balance; inflow (Receipt) increases it.
    opening_doc = await db.body_budgets.find_one(
        {"body_id": body_id, "fiscal_cycle": fiscal_cycle}, {"_id": 0},
    )
    opening = float((opening_doc or {}).get("opening_balance_inr", 0.0))

    rows: list[dict] = []
    running = opening
    total_dr = 0.0
    total_cr = 0.0
    for v in vouchers:
        # For the *owner* body_id (MPCA): Payment is Dr (grants given), Receipt is Cr (money in).
        # For a recipient division body: linked grants disbursed to them are Cr (money received).
        is_owner = v.get("body_id") == body_id
        vtype = v.get("voucher_type")
        amount = float(v.get("amount_inr") or 0)
        if is_owner:
            debit = amount if vtype == "Payment" else 0.0
            credit = amount if vtype == "Receipt" else 0.0
        else:
            # Recipient view: an outward payment from MPCA is money-in for the division
            debit = 0.0
            credit = amount if vtype == "Payment" else 0.0
        running += (credit - debit)
        total_dr += debit
        total_cr += credit
        rows.append({
            "date": v.get("date"),
            "voucher_no": v.get("voucher_no"),
            "voucher_type": vtype,
            "particulars": v.get("particulars"),
            "linked_ref_code": v.get("linked_ref_code"),
            "debit_inr": round(debit, 2),
            "credit_inr": round(credit, 2),
            "running_balance_inr": round(running, 2),
        })

    return {
        "body_id": body_id, "body_name": body["name"], "body_type": body["body_type"],
        "fiscal_cycle": fiscal_cycle,
        "opening_balance_inr": round(opening, 2),
        "rows": rows,
        "totals": {
            "debit_inr": round(total_dr, 2),
            "credit_inr": round(total_cr, 2),
            "closing_balance_inr": round(running, 2),
            "entry_count": len(rows),
        },
    }


# ═══════════════════ P3.6 · LEDGER ENDPOINT ═══════════════════

@api_router.get("/ledger")
async def ledger(body_id: str, fiscal_cycle: str | None = None,
                 date_from: str | None = None, date_to: str | None = None):
    fy = fiscal_cycle or indian_fy()
    return await _ledger_rows(body_id, fy, date_from, date_to)


# ═══════════════════ P3.7 · EXPORT UTILITY ═══════════════════

@api_router.get("/ledger/export.xlsx")
async def ledger_export_xlsx(body_id: str, fiscal_cycle: str | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fy = fiscal_cycle or indian_fy()
    data = await _ledger_rows(body_id, fy)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    # Title
    ws["A1"] = "MPCA · General Ledger"
    ws["A1"].font = Font(size=16, bold=True, color="0a1f3d")
    ws.merge_cells("A1:H1")
    ws["A2"] = f"{data['body_name']} ({data['body_id']}) · Fiscal {fy}"
    ws["A2"].font = Font(size=11, italic=True, color="7a1f2c")
    ws.merge_cells("A2:H2")

    # Opening
    ws["A4"] = "Opening Balance"
    ws["B4"] = data["opening_balance_inr"]
    ws["B4"].number_format = '₹ #,##0.00'

    # Headers
    headers = ["Date", "Voucher No.", "Type", "Particulars", "Ref Code", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
    header_fill = PatternFill("solid", fgColor="0a1f3d")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="7a1f2c")
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Rows
    for idx, r in enumerate(data["rows"], start=7):
        ws.cell(row=idx, column=1, value=r["date"])
        ws.cell(row=idx, column=2, value=r["voucher_no"])
        ws.cell(row=idx, column=3, value=r["voucher_type"])
        ws.cell(row=idx, column=4, value=r["particulars"])
        ws.cell(row=idx, column=5, value=r.get("linked_ref_code") or "")
        c_debit = ws.cell(row=idx, column=6, value=r["debit_inr"] or None)
        c_credit = ws.cell(row=idx, column=7, value=r["credit_inr"] or None)
        c_bal = ws.cell(row=idx, column=8, value=r["running_balance_inr"])
        for c in (c_debit, c_credit, c_bal):
            c.number_format = '#,##0.00'

    # Totals
    total_row = 7 + len(data["rows"])
    ws.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    ws.cell(row=total_row, column=6, value=data["totals"]["debit_inr"]).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=data["totals"]["credit_inr"]).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=data["totals"]["closing_balance_inr"]).font = Font(bold=True, color="0a1f3d")
    for col in (6, 7, 8):
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'

    # Column widths
    for col, width in enumerate([12, 20, 10, 40, 20, 14, 14, 16], start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    buf = BytesIO()
    await asyncio.to_thread(wb.save, buf)  # H4 · offload CPU-bound serialize
    buf.seek(0)
    filename = f"ledger_{body_id}_{fy}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/ledger/export.pdf")
async def ledger_export_pdf(body_id: str, fiscal_cycle: str | None = None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    fy = fiscal_cycle or indian_fy()
    data = await _ledger_rows(body_id, fy)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=12*mm, rightMargin=12*mm,
                             topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Title"], fontSize=16,
                               textColor=colors.HexColor("#0a1f3d"), alignment=1)
    sub_st = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10,
                             textColor=colors.HexColor("#7a1f2c"), alignment=1, italic=True)
    story = []
    story.append(Paragraph("MPCA · General Ledger", title_st))
    story.append(Paragraph(f"{data['body_name']} ({body_id}) · Fiscal {fy}", sub_st))
    story.append(Spacer(1, 6*mm))

    rows = [["Date", "Voucher No.", "Type", "Particulars", "Debit (₹)", "Credit (₹)", "Balance (₹)"]]
    rows.append(["", "", "", "Opening Balance", "", "", f"{data['opening_balance_inr']:,.2f}"])
    for r in data["rows"]:
        rows.append([
            r["date"], r["voucher_no"], r["voucher_type"],
            (r["particulars"][:60] + "…") if len(r["particulars"] or "") > 60 else r["particulars"],
            f"{r['debit_inr']:,.2f}" if r["debit_inr"] else "",
            f"{r['credit_inr']:,.2f}" if r["credit_inr"] else "",
            f"{r['running_balance_inr']:,.2f}",
        ])
    rows.append([
        "", "", "", "Totals",
        f"{data['totals']['debit_inr']:,.2f}",
        f"{data['totals']['credit_inr']:,.2f}",
        f"{data['totals']['closing_balance_inr']:,.2f}",
    ])
    tbl = Table(rows, colWidths=[22*mm, 42*mm, 22*mm, 90*mm, 28*mm, 28*mm, 32*mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0a1f3d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#7a1f2c")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#fbf7ed")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e9b949")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y · %H:%M UTC')}",
                            ParagraphStyle("foot", parent=styles["Normal"], fontSize=8,
                                           textColor=colors.grey, alignment=1)))
    await asyncio.to_thread(doc.build, story)  # H4 · offload CPU-bound PDF render
    buf.seek(0)
    filename = f"ledger_{body_id}_{fy}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ═══════════════════ P3.9 · BUDGET-VS-ACTUAL ═══════════════════

@api_router.get("/finance/budget-vs-actual")
async def budget_vs_actual(fiscal_cycle: str | None = None,
                            body_id: str | None = None):
    fy = fiscal_cycle or indian_fy()
    q_body: dict = {}
    if body_id: q_body["code"] = body_id
    else: q_body["body_type"] = {"$in": ["Division", "District"]}
    bodies = await db.bodies.find(q_body, {"_id": 0}).sort("code", 1).to_list(200)

    overrides = {b["body_id"]: b for b in await db.body_budgets.find(
        {"fiscal_cycle": fy}, {"_id": 0},
    ).to_list(500)}

    # Aggregate disbursed grants + vouchers
    dg_docs = await db.division_grants.find(
        {"fiscal_cycle": fy, "status": "Disbursed"}, {"_id": 0},
    ).to_list(2000)
    disbursed_by_body: dict = {}
    for g in dg_docs:
        amt = float(g.get("approved_amount_inr") or g.get("amount_inr") or 0)
        disbursed_by_body[g["body_id"]] = disbursed_by_body.get(g["body_id"], 0.0) + amt

    # Claims disbursed (older module still in use)
    claim_docs = await db.claims.find(
        {"fiscal_cycle": fy, "status": "Disbursed"}, {"_id": 0},
    ).to_list(2000)
    for c in claim_docs:
        amt = float(c.get("approved_amount_inr") or c.get("amount_inr") or 0)
        disbursed_by_body[c["body_id"]] = disbursed_by_body.get(c["body_id"], 0.0) + amt

    rows = []
    total_budget = 0.0
    total_actual = 0.0
    for b in bodies:
        code = b["code"]
        override = overrides.get(code)
        budget = float(override["annual_budget_inr"]) if override else float(b.get("annual_grant_inr") or 0)
        actual = round(disbursed_by_body.get(code, 0.0), 2)
        variance = round(budget - actual, 2)
        util = round((actual / budget) * 100, 1) if budget else 0.0
        status = "on_track" if util <= 100 else "over_budget"
        if 0 < util <= 60: status = "under_utilised"
        rows.append({
            "body_id": code,
            "body_name": b["name"],
            "body_type": b["body_type"],
            "annual_budget_inr": round(budget, 2),
            "actual_inr": actual,
            "variance_inr": variance,
            "utilisation_pct": util,
            "status": status,
        })
        total_budget += budget
        total_actual += actual

    return {
        "fiscal_cycle": fy,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_budget_inr": round(total_budget, 2),
        "total_actual_inr": round(total_actual, 2),
        "total_variance_inr": round(total_budget - total_actual, 2),
        "overall_utilisation_pct": round((total_actual / total_budget * 100), 1) if total_budget else 0.0,
        "rows": rows,
    }
